from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from datetime import datetime, date, timedelta, time as dtime
from difflib import SequenceMatcher
import os

from paratletismo_core.config.models import Sex, Category, EventType, FunctionalClassification
from paratletismo_core.tournaments.models import Institution, Athlete, Tournament, TournamentEvent
from paratletismo_core.competitions.models import Registration, AthleteEvent, Result, FinalResult

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = load_workbook = None

User = get_user_model()

ACCENTS = str.maketrans('áéíóúüñÁÉÍÓÚÜÑ', 'aeiouunAEIOUUN')


def norm(v):
    return str(v or '').translate(ACCENTS).lower()


def norm_key(v):
    import re
    return re.sub(r'[\s\-]+', '_', norm(v))


def norm_dni(v):
    return ''.join(ch for ch in str(v or '') if ch.isdigit())


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and v > 20000:
        from datetime import datetime as _dt
        return (_dt(1899, 12, 30) + timedelta(days=float(v))).date()
    if v is None or str(v).strip() in ('', '-'):
        return None
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y',
                '%d.%m.%Y', '%m/%d/%Y', '%m-%d-%Y', '%Y%m%d',
                '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_datetime(v, fallback_time=9):
    d = to_date(v)
    if d is None:
        return None
    dt = datetime.combine(d, dtime(fallback_time))
    return timezone.make_aware(dt) if timezone.is_naive(dt) else dt


def parse_mark(mark):
    if mark is None or str(mark).strip() == '':
        return None
    try:
        return float(str(mark).replace(',', '.'))
    except ValueError:
        pass
    parts = str(mark).replace(',', '.').split(':')
    try:
        if len(parts) == 2:
            return float(parts[0]) * 60 + float(parts[1])
        if len(parts) == 3:
            return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
    except ValueError:
        pass
    return None


def parse_wind(v):
    """Convierte el viento a un numero (m/s). Vacio/no valido devuelve None."""
    if v is None or str(v).strip() in ('', '-'):
        return None
    try:
        return round(float(str(v).replace(',', '.').replace('m/s', '').replace('m', '').strip()), 1)
    except ValueError:
        return None


def resolve_sex(label):
    if label is None:
        return None
    key = norm_key(label)
    if key in ('m', 'masculino', 'men'):
        return Sex.objects.filter(code__iexact='M').first() or Sex.objects.filter(name__iexact='Masculino').first()
    if key in ('f', 'femenino', 'women', 'fem'):
        return Sex.objects.filter(code__iexact='F').first() or Sex.objects.filter(name__iexact='Femenino').first()
    return Sex.objects.filter(code__iexact=norm_key(label)).first() or Sex.objects.filter(name__iexact=str(label)).first()


def resolve_classification(label):
    if label is None:
        return None
    key = norm_key(label)
    ex = FunctionalClassification.objects.filter(code__iexact=key).first()
    if ex:
        return ex
    if key.isdigit():
        for c in FunctionalClassification.objects.all():
            if c.code and c.code[-len(key):] == key:
                return c
    return None


def resolve_category(label):
    if label is None:
        return None
    ex = Category.objects.filter(name__iexact=str(label)).first()
    if ex:
        return ex
    k = _plural_norm(str(label))
    for c in Category.objects.all():
        if k and _plural_norm(c.name) == k:
            return c
    return None


def _plural_norm(v):
    s = _norm_event(v)
    if s.endswith('es') and len(s) > 4:
        return s[:-2]
    if s.endswith('s') and len(s) > 3:
        return s[:-1]
    return s


def _norm_event(v):
    return ''.join(ch for ch in norm(v) if ch.isalnum())


EVENT_TYPE_ALIASES = {
    'slargo': 'Salto Largo',
    'saltoenlargo': 'Salto Largo',
    'saltoenlongitud': 'Salto Largo',
    'lbala': 'Lanzamiento de bala',
    'lanzamientobala': 'Lanzamiento de bala',
    'bala': 'Lanzamiento de bala',
    'lclava': 'Lanzamiento Clava',
    'clava': 'Lanzamiento Clava',
    'ldisco': 'Lanzamiento de disco',
    'disco': 'Lanzamiento de disco',
    'ljabalina': 'Lanzamiento de jabalina',
    'jabalina': 'Lanzamiento de jabalina',
    'lmartillo': 'Lanzamiento de martillo',
    'martillo': 'Lanzamiento de martillo',
    'saltoalto': 'Salto Alto',
    'relevo4x100m': 'Relevo 4x100m',
    '4x100m': 'Relevo 4x100m',
}


def resolve_event_type(label):
    if label is None:
        return None
    ex = EventType.objects.filter(name__iexact=str(label)).first()
    if ex:
        return ex
    k = _norm_event(str(label))
    for et in EventType.objects.all():
        if _norm_event(et.name) == k:
            return et
    alias = EVENT_TYPE_ALIASES.get(k)
    if alias:
        return EventType.objects.filter(name__iexact=alias).first()
    return None


ESTADO_MAP = {
    'completado': 'completed',
    'completed': 'completed',
    'en_progreso': 'in_progress',
    'inscripcion_abierta': 'registration_open',
    'inscripcion_cerrada': 'registration_closed',
    'borrador': 'draft',
    'cancelado': 'cancelled',
    'cancelled': 'cancelled',
}


class Command(BaseCommand):
    help = ('Importa datos historicos desde un archivo Excel (xlsx). '
            'Hojas: Instituciones, Atletas, Torneos, Pruebas, Resultados. '
            'Usa --crear-plantilla para generar la plantilla de ejemplo.')

    def add_arguments(self, parser):
        parser.add_argument('--file', dest='file')
        parser.add_argument('--url', dest='url',
                            help='Descarga la planilla desde una URL antes de importar '
                                 '(util para produccion, ej. un link publico del xlsx).')
        parser.add_argument('--crear-plantilla', dest='plantilla', default='')
        parser.add_argument('--admin-email', dest='admin_email', default='')
        parser.add_argument('--password', dest='password', default='atleta123')

    def handle(self, *args, **opts):
        if Workbook is None:
            raise CommandError('openpyxl no esta instalado. Ejecuta: pip install openpyxl')

        if opts.get('plantilla'):
            self._crear_plantilla(opts['plantilla'])
            return

        if not opts.get('file'):
            if not opts.get('url'):
                raise CommandError('Debes indicar --file ruta/al/archivo.xlsx (o --crear-plantilla ruta)')
            import tempfile
            import urllib.request
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.close()
            try:
                self.stdout.write(f'Descargando planilla desde: {opts["url"]}')
                urllib.request.urlretrieve(opts['url'], tmp.name)
            except Exception as exc:
                os.remove(tmp.name)
                raise CommandError(f'No se pudo descargar la planilla: {exc}')
            opts['file'] = tmp.name
        if not self._existe(opts['file']):
            raise CommandError(f'No se encontro el archivo: {opts["file"]}')

        admin = None
        if opts.get('admin_email'):
            admin = User.objects.filter(email=opts['admin_email']).first()
        if admin is None:
            admin = User.objects.filter(is_superuser=True).order_by('date_joined').first()
        if admin is None:
            raise CommandError('No hay ningun superusuario. Ejecuta primero: python manage.py seed_data')

        self.password = opts['password']

        try:
            wb = load_workbook(opts['file'], data_only=True)
        except Exception as e:
            raise CommandError(f'No se pudo abrir el Excel: {e}')

        with transaction.atomic():
            self._importar(wb, admin)

        self.stdout.write(self.style.SUCCESS('\nImportacion finalizada.'))

    # ---------- utilidades ----------

    def _existe(self, path):
        from django.conf import settings
        import os
        if os.path.exists(path):
            return True
        return os.path.exists(os.path.join(settings.BASE_DIR, path))

    def _abrir_ruta(self, path):
        from django.conf import settings
        import os
        if os.path.exists(path):
            return path
        return os.path.join(settings.BASE_DIR, path)

    def _hoja(self, wb, nombre):
        for ws in wb.worksheets:
            if norm_key(ws.title) == norm_key(nombre):
                return ws
        return None

    def _filas(self, ws):
        if ws is None:
            return []
        rows = []
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                values = [c for c in row]
                if not any(v not in (None, '') for v in values):
                    continue
                header = [norm_key(v) for v in values]
                continue
            if not any(v not in (None, '') for v in row):
                continue
            rows.append(dict(zip(header, row)))
        return rows

    def _dato(self, fila, *keys):
        for k in keys:
            v = fila.get(norm_key(k))
            if v not in (None, ''):
                return v
        return None

    def _resolver_clasificacion(self, label):
        """Devuelve (track_fc, field_fc, generic_fc) para asignar al atleta.

        La planilla puede traer clasificacion solo con el numero (ej '20') o con
        letra ('T20'/'F20'). Cuando viene solo el numero no se sabe si es de pista
        (T) o campo (F), asi que se completan AMBAS clasificaciones (track y field)
        para que la pagina publica resuelva segun el tipo de prueba. El campo
        generico se completa solo si la letra es explicita."""
        label = str(label or '').strip().upper()
        if not label:
            return (None, None, None)
        import re as _re
        m = _re.fullmatch(r'([TF])?(\d+)', label)
        prefijo = m.group(1) if m else None
        numero = m.group(2) if m else None
        def _por_codigo(code):
            return FunctionalClassification.objects.filter(code__iexact=code).first()
        if prefijo:
            exact = _por_codigo(label)
            if prefijo == 'T':
                return (exact, None, exact)
            return (None, exact, exact)
        if numero:
            track_fc = _por_codigo('T' + numero)
            field_fc = _por_codigo('F' + numero)
            return (track_fc, field_fc, None)
        return (None, None, None)

    def _match_institucion(self, nombre, instituciones, instituciones_sigla):
        texto = str(nombre or '').strip()
        if not texto or norm_key(texto) in ('sin_institucion', 'sin_datos', 'sin_datos_para_mostrar', 'ninguna', '_', 's_d', 'sd', 's_n'):
            return None
        clave = norm_key(texto)
        inst = instituciones.get(clave)
        if inst:
            return inst
        inst = instituciones_sigla.get(clave)
        if inst:
            return inst
        if len(clave) < 4:
            return None
        mejor = (0.0, None, '')
        for ck, cand in list(instituciones.items()) + list(instituciones_sigla.items()):
            if len(ck) < 4:
                continue
            if clave in ck or ck in clave:
                return cand
            r = SequenceMatcher(None, clave, ck).ratio()
            if r > mejor[0]:
                mejor = (r, cand, ck)
        ratio, cand, ck = mejor
        if ratio >= 0.78 and len(set(clave.split('_')) & set(ck.split('_'))) >= 2:
            return cand
        return None

    def _reporte(self, nombre, creados, omitidos, errores):
        self.stdout.write(f'\n[{nombre}] creados: {creados}, existentes (no duplicados): {omitidos}, errores: {len(errores)}')
        for e in errores[:20]:
            self.stdout.write(self.style.WARNING(f'  - {e}'))

    # ---------- importacion ----------

    def _importar(self, wb, admin):
        self.stdout.write(self.style.SUCCESS(f'Importando con admin/organizador: {admin.email}'))

        # 1. instituciones
        instituciones = {}
        creadas = omit = 0
        errores = []
        for fila in self._filas(self._hoja(wb, 'instituciones')):
            nombre = self._dato(fila, 'nombre')
            if not nombre:
                errores.append('Institucion sin nombre (fila ignorada)')
                continue
            inst, created = Institution.objects.get_or_create(
                name=str(nombre),
                defaults={
                    'short_name': self._dato(fila, 'sigla') or '',
                    'city': self._dato(fila, 'ciudad') or '',
                    'province': self._dato(fila, 'provincia') or '',
                    'phone': self._dato(fila, 'telefono') or '',
                    'email': self._dato(fila, 'email') or '',
                    'is_active': True,
                    'can_organize': False,
                })
            instituciones[norm_key(inst.name)] = inst
            if created:
                creadas += 1
            else:
                omit += 1
        self._reporte('Instituciones', creadas, omit, errores)

        # indexar TODAS las instituciones del sistema (por nombre y sigla): asi el organizador
        # puede cargarse como "coparg" y matchear "Comite Paralimpico" aunque no este en el sheet
        instituciones_sigla = {}
        for inst in Institution.objects.all():
            instituciones.setdefault(norm_key(inst.name), inst)
            if inst.short_name:
                instituciones_sigla.setdefault(norm_key(inst.short_name), inst)

        # 2. atletas (identificados por DNI/documento, unico por atleta)
        atletas_por_dni = {}
        atletas_por_email = {}
        emails_usados = set()
        for u in User.objects.values_list('email', flat=True):
            if u:
                emails_usados.add(norm_key(u))
        documentos = {}
        creados = omit = 0
        errores = []
        for fila in self._filas(self._hoja(wb, 'atletas')):
            nombre = self._dato(fila, 'nombre')
            apellido = self._dato(fila, 'apellido')
            doc = str(self._dato(fila, 'documento') or '').strip()
            if not nombre or not apellido or not doc:
                errores.append('Atleta incompleto (nombre, apellido y documento/DNI obligatorios)')
                continue
            clave_dni = norm_dni(doc)
            if clave_dni in documentos:
                errores.append(f'Documento repetido "{doc}": ya usado por {documentos[clave_dni]}. Fila ignorada.')
                continue
            if Athlete.objects.filter(document_number=doc).exists():
                athlete = Athlete.objects.get(document_number=doc)
                # Se identifican por DNI. Cuando el atleta YA existe, solo se CORRIGE la
                # identificacion (nombre/apellido) que vinieron mal escritos en la planilla.
                # El resto de campos (fecha de nacimiento, telefono, clasificacion, sexo,
                # institucion, etc.) NO se tocan para no pisar datos sensibles o cargados a mano.
                # El nombre/apellido se guarda en el User asociado al atleta, no en Athlete.
                actualizado = False
                if athlete.user:
                    for fld, val in (('first_name', str(nombre)), ('last_name', str(apellido))):
                        v = val.strip()
                        if v and getattr(athlete.user, fld) != v:
                            setattr(athlete.user, fld, v)
                            actualizado = True
                    if actualizado:
                        athlete.user.save()
                atletas_por_dni[clave_dni] = athlete
                atletas_por_email[norm_key(athlete.user.email)] = athlete
                emails_usados.add(norm_key(athlete.user.email))
                documentos[clave_dni] = f'{nombre} {apellido}'
                omit += 1
                continue

            # el email es opcional: si falta (o se repite) se genera uno a partir del DNI
            email = str(self._dato(fila, 'email') or '').strip()
            email_usado = norm_key(email) if email else ''
            if not email or email_usado in emails_usados:
                email = f'dni{doc}@import.local'
                email_usado = norm_key(email)
                if email_usado in emails_usados:
                    email = f'dni{doc}.{len(emails_usados) + 1}@import.local'
                    email_usado = norm_key(email)
            user = User.objects.create_user(
                email=email,
                password=self.password,
                first_name=str(nombre),
                last_name=str(apellido),
                phone=self._dato(fila, 'telefono') or '',
                role='athlete',
            )
            emails_usados.add(email_usado)
            inst_name = self._dato(fila, 'institucion')
            institution = self._match_institucion(inst_name, instituciones, instituciones_sigla)
            sexo = resolve_sex(self._dato(fila, 'sexo'))
            track_fc, field_fc, generic_fc = self._resolver_clasificacion(self._dato(fila, 'clasificacion'))
            athlete = Athlete.objects.create(
                user=user,
                institution=institution,
                document_type='dni',
                document_number=doc,
                date_of_birth=to_date(self._dato(fila, 'fecha_nacimiento')),
                sex=sexo,
                functional_classification=generic_fc,
                track_classification=track_fc,
                field_classification=field_fc,
                phone=self._dato(fila, 'telefono') or '',
            )
            atletas_por_dni[clave_dni] = athlete
            atletas_por_email[email_usado] = athlete
            documentos[clave_dni] = f'{nombre} {apellido}'
            creados += 1
        self._reporte('Atletas', creados, omit, errores)

        # 3. torneos
        torneos = {}
        creados = omit = 0
        errores = []
        for fila in self._filas(self._hoja(wb, 'torneos')):
            nombre = self._dato(fila, 'nombre')
            if not nombre:
                errores.append('Torneo sin nombre')
                continue
            fecha_ini = to_datetime(self._dato(fila, 'fecha_inicio'), 9)
            fecha_fin = to_datetime(self._dato(fila, 'fecha_fin'), 19)
            if not fecha_ini:
                fecha_ini = timezone.now()
            if not fecha_fin:
                fecha_fin = fecha_ini + timedelta(days=1)
            if Tournament.objects.filter(name=str(nombre)).exists():
                t0 = Tournament.objects.get(name=str(nombre))
                # solo se actualizan las fechas si la planilla las trae (comportamiento original)
                if self._dato(fila, 'fecha_inicio') not in (None, ''):
                    Tournament.objects.filter(id=t0.id).update(
                        tournament_start=fecha_ini,
                        tournament_end=fecha_fin,
                        registration_opens=fecha_ini - timedelta(days=20),
                        registration_closes=fecha_ini - timedelta(days=2),
                    )
                torneos[norm_key(nombre)] = t0
                omit += 1
                continue
            org_name = self._dato(fila, 'organizador')
            organizer = self._match_institucion(org_name, instituciones, instituciones_sigla) if org_name else None
            estado = norm_key(self._dato(fila, 'estado') or 'completado')
            tournament = Tournament(
                name=str(nombre),
                description=self._dato(fila, 'descripcion') or '',
                organizer=organizer,
                admin_user=admin,
                venue=self._dato(fila, 'sede') or '',
                address=self._dato(fila, 'direccion') or '',
                city=self._dato(fila, 'ciudad') or '',
                province=self._dato(fila, 'provincia') or '',
                status=ESTADO_MAP.get(estado, 'completed'),
                payment_status='paid',
                payment_amount=None,
                payment_date=timezone.now(),
                paid_by=admin,
                is_active=True,
                registration_opens=fecha_ini - timedelta(days=20),
                registration_closes=fecha_ini - timedelta(days=2),
                tournament_start=fecha_ini,
                tournament_end=fecha_fin,
                registration_fee=0,
            )
            tournament.save()
            torneos[norm_key(nombre)] = tournament
            creados += 1
        self._reporte('Torneos', creados, omit, errores)

        # 4. pruebas
        eventos = {}
        vientos = {}
        creados = omit = 0
        errores = []
        for fila in self._filas(self._hoja(wb, 'pruebas')):
            nombre_t = self._dato(fila, 'torneo')
            nombre_p = self._dato(fila, 'nombre')
            tournament = torneos.get(norm_key(nombre_t)) if nombre_t else None
            if tournament is None:
                errores.append(f'Prueba "{nombre_p}": torneo no encontrado')
                continue
            viento_prueba = parse_wind(self._dato(fila, 'viento'))
            if TournamentEvent.objects.filter(tournament=tournament, name=str(nombre_p)).exists():
                event = TournamentEvent.objects.get(tournament=tournament, name=str(nombre_p))
                eventos[(tournament.id, norm_key(nombre_p))] = event
                if viento_prueba is not None:
                    vientos[(tournament.id, norm_key(nombre_p))] = viento_prueba
                omit += 1
                continue
            tipo = resolve_event_type(self._dato(fila, 'tipo'))
            if tipo is None:
                errores.append(f'Prueba "{nombre_p}": tipo de prueba no encontrado ("{self._dato(fila, "tipo")}")')
                continue
            event = TournamentEvent(
                tournament=tournament,
                name=str(nombre_p),
                event_type=tipo,
                discipline=tipo.discipline,
                sex=resolve_sex(self._dato(fila, 'sexo')),
                category=resolve_category(self._dato(fila, 'categoria')),
                functional_classification=resolve_classification(self._dato(fila, 'clasificacion')),
                is_final=True,
                status='completed',
            )
            event.save()
            eventos[(tournament.id, norm_key(nombre_p))] = event
            if viento_prueba is not None:
                vientos[(tournament.id, norm_key(nombre_p))] = viento_prueba
            creados += 1
        self._reporte('Pruebas', creados, omit, errores)

        # poblar m2m de cada torneo segun sus pruebas
        for tournament in Tournament.objects.filter(id__in=[t.id for t in torneos.values()]):
            evs = tournament.events.all()
            tournament.disciplines.set(list(set(e.discipline for e in evs if e.discipline)))
            tournament.sexes.set(list(set(e.sex for e in evs if e.sex)))
            tournament.categories.set(list(set(e.category for e in evs if e.category)))
            tournament.functional_classifications.set(
                list(set(e.functional_classification for e in evs if e.functional_classification)))

        # 5. resultados (crea registrations, athlete_events, results y final_results)
        creados = omit = 0
        errores = []
        registrados = 0
        for fila in self._filas(self._hoja(wb, 'resultados')):
            nombre_t = self._dato(fila, 'torneo')
            nombre_p = self._dato(fila, 'prueba')
            dni = self._dato(fila, 'atleta_dni') or self._dato(fila, 'dni')
            email = self._dato(fila, 'atleta_email')
            if not nombre_t or not nombre_p or not dni and not email:
                errores.append('Resultado incompleto (torneo, prueba y atleta_dni obligatorios)')
                continue
            tournament = torneos.get(norm_key(nombre_t))
            if tournament is None:
                errores.append(f'Resultado: torneo no encontrado "{nombre_t}"')
                continue
            event = eventos.get((tournament.id, norm_key(nombre_p)))
            if event is None:
                errores.append(f'Resultado: prueba no encontrada "{nombre_p}" en torneo "{nombre_t}"')
                continue
            athlete = atletas_por_dni.get(norm_dni(dni)) if dni else None
            if athlete is None and email:
                athlete = atletas_por_email.get(norm_key(email))
            if athlete is None:
                ref = dni or email
                errores.append(f'Resultado: atleta no encontrado "{ref}" (revisar hoja Atletas)')
                continue
            if athlete.institution is None:
                athlete.institution = tournament.organizer
                athlete.save()

            registration, r_created = Registration.objects.get_or_create(
                tournament=tournament,
                athlete=athlete,
                defaults={
                    'institution': athlete.institution,
                    'status': 'approved',
                    'payment_status': 'exempt',
                })
            if r_created:
                registrados += 1

            ae, ae_created = AthleteEvent.objects.get_or_create(
                registration=registration,
                tournament_event=event,
                defaults={'status': 'confirmed'})

            marca = self._dato(fila, 'marca')
            valor = parse_mark(marca)
            viento = parse_wind(self._dato(fila, 'viento'))
            if viento is None:
                viento = vientos.get((tournament.id, norm_key(nombre_p)))
            Result.objects.update_or_create(
                athlete_event=ae,
                attempt_number=1,
                defaults={
                    'mark': str(marca) if marca else '',
                    'value': valor,
                    'is_valid': True,
                    'wind': viento,
                })

            rank = self._dato(fila, 'rank')
            if rank is not None and str(rank).strip() != '':
                try:
                    rank = int(float(str(rank)))
                except ValueError:
                    rank = None
            if rank is not None:
                FinalResult.objects.update_or_create(
                    tournament_event=event,
                    athlete=athlete,
                    defaults={
                        'rank': rank,
                        'best_mark': str(marca) if marca else '',
                    })
                creados += 1
            else:
                omit += 1
        self._reporte('Resultados (FinalResults)', creados, omit, errores)
        self.stdout.write(f'   Inscripciones (Registrations) creadas: {registrados}')

    # ---------- plantilla ----------

    def _crear_plantilla(self, ruta):
        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2F5597')

        def write_sheet(ws, headers, rows, widths):
            ws.append(headers)
            for row in rows:
                ws.append(row)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = 'A2'

        instrucciones = [
            ['IMPORTADOR DE DATOS HISTORICOS'],
            [''],
            ['Completa las hojas y despues ejecuta:'],
            ['  python manage.py import_legacy --file C:\\ruta\\archivo.xlsx'],
            [''],
            ['Para ver esta plantilla con datos de ejemplo cargados, ejecuta:'],
            ['  python manage.py import_legacy --crear-plantilla ruta.xlsx'],
            [''],
            ['VALIDACIONES'],
            ['- Las hojas se buscan por nombre (el orden no importa).'],
            ['- Atleta se identifica por DOCUMENTO/DNI (unico por atleta). Si el DNI ya existe,'],
            ['  se CORRIGE solo el nombre/apellido mal escritos de la planilla; los demas datos'],
            ['  (fecha nac, telefono, clasificacion, sexo, institucion) NO se modifican.'],
            ['  Asi podes corregir nombres equivocados o apellidos duplicados en el Excel y'],
            ['  propagarlos a la web al volver a importar.'],
            ['- Torneo si ya existe: se omite (no se duplica). Se actualizan fechas si vienen.'],
            ['- Instituto si ya existe: se omite (no se duplica). Nuevo entra DESHABILITADO.'],
            ['- EMAIL OPCIONAL en Atletas: si lo dejas vacio el sistema genera uno automatico a partir del DNI'],
            ['  (ej: dni30123456@import.local). Asi los menores que no tienen correo propio no necesitan uno:'],
            ['  pueden compartir el email del padre o dejarlo vacio; el sistema nunca mezcla atletas porque'],
            ['  la identificacion es por DNI.'],
            ['- Referencias por nombre exacto: tipo de prueba (ej: "100 mts"), categoria (ej: "Mayor"),'],
            ['  clasificacion funcional (ej: "T12" o "F57"), sexo ("Masculino"/"Femenino").'],
            ['- INSTITUCIONES: opcionales y nunca generan error. Se buscan por nombre o sigla, sin acentos y'],
            ['  tolerando variaciones del nombre (ej: "Estrella del Sur" = "Club Atletico Estrella del Sur").'],
            ['  Si falta la columna, esta vacia o el nombre no se encuentra, el atleta/torneo queda SIN'],
            ['  institucion (sin error). En ATLETAS la institucion solo se asigna si el atleta ya no tiene una.'],
            ['- El ORGANIZADOR del torneo puede escribirse por nombre o por sigla (ej: "coparg" = "Comite Paralimpico").'],
            ['- Si un torneo no aparece en la hoja Pruebas no se le agregan disciplinas/sexos/categorias.'],
            ['- Fechas en formato AAAA-MM-DD (ej: 2024-11-15).'],
            [''],
            ['VIENTO (en m/s)'],
            ['- Columna "viento" en la hoja PRUEBAS: viento unico para carreras (ej: 100 mts), igual para todos los'],
            ['  atletas de esa prueba. Ej: +1.2 (viento a favor), -0.4 (viento en contra).'],
            ['- Columna "viento" en la hoja RESULTADOS: viento INDIVIDUAL por atleta (ej: salto en largo, se mide'],
            ['  en cada salto). Si la dejas vacia se usa el viento cargado en la hoja Pruebas.'],
            ['- Si la columna queda vacia y no hay viento en Pruebas, no se registra viento.'],
        ]
        ini = wb.active
        ini.title = 'Instrucciones'
        write_sheet(ini, ['Instrucciones'], instrucciones, [70])

        instituciones = [
            ['nombre', 'sigla', 'ciudad', 'provincia', 'telefono', 'email'],
            ['Club Atletico Estrella del Sur', 'CES', 'Mar del Plata', 'Buenos Aires', '0223-1234567', 'contacto@estrelladelsur.com'],
            ['Federacion Deportiva Provinciana', 'FDP', 'Rosario', 'Santa Fe', '0341-7654321', 'info@fdp.org.ar'],
            ['Fundacion Aguilas Blancas', 'FAB', 'Cordoba', 'Cordoba', '0351-9876543', 'aguilasblancas@gmail.com'],
        ]
        write_sheet(wb.create_sheet('Instituciones'), instituciones[0], instituciones[1:], [34, 10, 20, 16, 18, 34])

        atletas = [
            ['documento', 'email', 'nombre', 'apellido', 'fecha_nacimiento', 'sexo', 'clasificacion', 'institucion'],
            ['30123456', 'juan.perez@demo.com', 'Juan', 'Perez', '1998-06-15', 'Masculino', 'T12', 'Club Atletico Estrella del Sur'],
            ['30678901', 'luis.romero@demo.com', 'Luis', 'Romero', '1997-05-25', 'Masculino', 'T12', 'Club Atletico Estrella del Sur'],
            ['31123456', 'camila.vega@demo.com', 'Camila', 'Vega', '1998-02-28', 'Femenino', 'T12', 'Federacion Deportiva Provinciana'],
            ['30234567', 'maria.gomez@demo.com', 'Maria', 'Gomez', '1996-03-22', 'Femenino', 'T37', 'Club Atletico Estrella del Sur'],
            ['31345678', 'beatriz.castro@demo.com', 'Beatriz', 'Castro', '1999-07-14', 'Femenino', 'T37', 'Federacion Deportiva Provinciana'],
            ['30789012', 'sofia.diaz@demo.com', 'Sofia', 'Diaz', '1995-09-17', 'Femenino', 'T38', 'Fundacion Aguilas Blancas'],
            ['31456789', 'paulina.rojas@demo.com', 'Paulina', 'Rojas', '2001-01-09', 'Femenino', 'T38', 'Fundacion Aguilas Blancas'],
            ['30345678', 'carlos.ruiz@demo.com', 'Carlos', 'Ruiz', '2001-11-02', 'Masculino', 'F57', 'Federacion Deportiva Provinciana'],
            ['31567890', 'rodrigo.silva@demo.com', 'Rodrigo', 'Silva', '1994-12-03', 'Masculino', 'F57', 'Federacion Deportiva Provinciana'],
            ['30456789', 'ana.fernandez@demo.com', 'Ana', 'Fernandez', '1999-08-30', 'Femenino', 'F57', 'Fundacion Aguilas Blancas'],
            ['31678901', 'marisa.cabrera@demo.com', 'Marisa', 'Cabrera', '1996-04-11', 'Femenino', 'F57', 'Fundacion Aguilas Blancas'],
            ['30567890', 'pedro.sosa@demo.com', 'Pedro', 'Sosa', '2000-01-10', 'Masculino', 'T20', 'Fundacion Aguilas Blancas'],
            ['31234567', 'diego.acosta@demo.com', 'Diego', 'Acosta', '2000-10-19', 'Masculino', 'T20', 'Club Atletico Estrella del Sur'],
            ['30901234', 'valentina.arias@demo.com', 'Valentina', 'Arias', '2004-12-12', 'Femenino', 'F64', 'Club Atletico Estrella del Sur'],
            ['31789012', 'julieta.nunez@demo.com', 'Julieta', 'Nunez', '1997-08-08', 'Femenino', 'F64', 'Federacion Deportiva Provinciana'],
            ['31012345', 'jorge.morales@demo.com', 'Jorge', 'Morales', '1993-07-07', 'Masculino', 'F64', 'Fundacion Aguilas Blancas'],
            ['31290123', 'alejandro.mendez@demo.com', 'Alejandro', 'Mendez', '1995-03-29', 'Masculino', 'F64', 'Fundacion Aguilas Blancas'],
            ['31401234', 'luciana.robles@demo.com', 'Luciana', 'Robles', '2002-06-21', 'Femenino', 'T12', 'Club Atletico Estrella del Sur'],
        ]
        write_sheet(wb.create_sheet('Atletas'), atletas[0], atletas[1:], [14, 30, 15, 15, 18, 12, 14, 30])

        torneos = [
            ['nombre', 'descripcion', 'sede', 'direccion', 'ciudad', 'provincia', 'organizador', 'fecha_inicio', 'fecha_fin', 'estado'],
            ['Campeonato Provincial de Paratletismo 2024', 'Competencia provincial de atletismo adaptado', 'Estadio Municipal', 'Av. Del Deporte 1200', 'Rosario', 'Santa Fe', 'Federacion Deportiva Provinciana', '2024-11-15', '2024-11-17', 'completado'],
            ['Gran Prix Nacional de Paratletismo 2025', 'Gran premio nacional de atletismo paralimpico', 'Polideportivo Barrio Norte', 'Calle 12 2300', 'Mar del Plata', 'Buenos Aires', 'Club Atletico Estrella del Sur', '2025-05-09', '2025-05-11', 'completado'],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Torneo interprovincial organizado por Aguilas Blancas', 'Complejo Chateau', 'Av. Bicentenario 500', 'Cordoba', 'Cordoba', 'Fundacion Aguilas Blancas', '2023-10-20', '2023-10-22', 'completado'],
        ]
        write_sheet(wb.create_sheet('Torneos'), torneos[0], torneos[1:], [42, 42, 30, 24, 16, 16, 30, 14, 14, 14])

        pruebas = [
            ['torneo', 'nombre', 'tipo', 'sexo', 'categoria', 'clasificacion', 'viento'],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Masculino Mayor T12', '100 mts', 'Masculino', 'Mayor', 'T12', '+1.2'],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Femenino Mayor T37', '100 mts', 'Femenino', 'Mayor', 'T37', '-0.4'],
            ['Campeonato Provincial de Paratletismo 2024', 'Bala Masculino Mayor F57', 'Lanzamiento de bala', 'Masculino', 'Mayor', 'F57', ''],
            ['Campeonato Provincial de Paratletismo 2024', 'Salto en longitud Femenino Mayor F64', 'Salto Largo', 'Femenino', 'Mayor', 'F64', ''],
            ['Campeonato Provincial de Paratletismo 2024', '1500 mts Masculino Mayor T20', '1500 mts', 'Masculino', 'Mayor', 'T20', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '200 mts Masculino Mayor T12', '200 mts', 'Masculino', 'Mayor', 'T12', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '100 mts Femenino Mayor T38', '100 mts', 'Femenino', 'Mayor', 'T38', '+0.8'],
            ['Gran Prix Nacional de Paratletismo 2025', 'Bala Femenino Mayor F57', 'Lanzamiento de bala', 'Femenino', 'Mayor', 'F57', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Disco Masculino Mayor F57', 'Lanzamiento de disco', 'Masculino', 'Mayor', 'F57', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Salto en longitud Masculino Mayor F64', 'Salto Largo', 'Masculino', 'Mayor', 'F64', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Masculino Mayor T20', '100 mts', 'Masculino', 'Mayor', 'T20', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Femenino Mayor T12', '100 mts', 'Femenino', 'Mayor', 'T12', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Bala Masculino Mayor F57', 'Lanzamiento de bala', 'Masculino', 'Mayor', 'F57', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Salto en longitud Femenino Mayor F64', 'Salto Largo', 'Femenino', 'Mayor', 'F64', ''],
        ]
        write_sheet(wb.create_sheet('Pruebas'), pruebas[0], pruebas[1:], [44, 38, 24, 14, 12, 16, 10])

        resultados = [
            ['torneo', 'prueba', 'atleta_dni', 'marca', 'rank', 'viento'],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Masculino Mayor T12', '30123456', '11.88', '1', ''],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Masculino Mayor T12', '30678901', '12.47', '2', ''],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Femenino Mayor T37', '30234567', '14.05', '1', ''],
            ['Campeonato Provincial de Paratletismo 2024', '100 mts Femenino Mayor T37', '31345678', '14.66', '2', ''],
            ['Campeonato Provincial de Paratletismo 2024', 'Bala Masculino Mayor F57', '30345678', '9.42', '1', ''],
            ['Campeonato Provincial de Paratletismo 2024', 'Bala Masculino Mayor F57', '31567890', '8.87', '2', ''],
            ['Campeonato Provincial de Paratletismo 2024', 'Salto en longitud Femenino Mayor F64', '30901234', '4.78', '1', '+0.6'],
            ['Campeonato Provincial de Paratletismo 2024', 'Salto en longitud Femenino Mayor F64', '31789012', '4.35', '2', '-1.1'],
            ['Campeonato Provincial de Paratletismo 2024', '1500 mts Masculino Mayor T20', '30567890', '4:32.10', '1', ''],
            ['Campeonato Provincial de Paratletismo 2024', '1500 mts Masculino Mayor T20', '31234567', '4:40.55', '2', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '200 mts Masculino Mayor T12', '30678901', '24.12', '1', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '200 mts Masculino Mayor T12', '30123456', '24.85', '2', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '100 mts Femenino Mayor T38', '30789012', '13.91', '1', ''],
            ['Gran Prix Nacional de Paratletismo 2025', '100 mts Femenino Mayor T38', '31456789', '14.30', '2', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Bala Femenino Mayor F57', '30456789', '8.02', '1', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Bala Femenino Mayor F57', '31678901', '7.45', '2', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Disco Masculino Mayor F57', '31567890', '24.10', '1', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Disco Masculino Mayor F57', '30345678', '22.55', '2', ''],
            ['Gran Prix Nacional de Paratletismo 2025', 'Salto en longitud Masculino Mayor F64', '31012345', '5.12', '1', '+1.3'],
            ['Gran Prix Nacional de Paratletismo 2025', 'Salto en longitud Masculino Mayor F64', '31290123', '4.86', '2', '+0.2'],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Masculino Mayor T20', '30567890', '11.62', '1', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Masculino Mayor T20', '31234567', '12.01', '2', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Femenino Mayor T12', '31123456', '13.75', '1', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', '100 mts Femenino Mayor T12', '31401234', '14.10', '2', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Bala Masculino Mayor F57', '31567890', '9.85', '1', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Bala Masculino Mayor F57', '30345678', '9.10', '2', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Salto en longitud Femenino Mayor F64', '30901234', '4.95', '1', ''],
            ['Torneo Interprovincial Aguilas del Sur 2023', 'Salto en longitud Femenino Mayor F64', '31789012', '4.52', '2', ''],
        ]
        write_sheet(wb.create_sheet('Resultados'), resultados[0], resultados[1:], [44, 38, 14, 14, 10, 10])

        wb.save(ruta)
        self.stdout.write(self.style.SUCCESS(f'Plantilla creada en: {ruta}'))
        self.stdout.write('Completa las hojas y luego ejecuta: python manage.py import_legacy --file <ruta>')